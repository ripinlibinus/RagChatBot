# file: eval_with_excel_audit.py
import os
import re
import json, ast
import unicodedata
import requests
from datetime import datetime
from rapidfuzz import fuzz
from sklearn.metrics import precision_recall_fscore_support
from read_query_xls import build_answer_gold_list

import pandas as pd
from openpyxl import load_workbook

from dotenv import load_dotenv
load_dotenv()

# ===============================
# 0) KONFIG OUTPUT AUDIT
# ===============================
AUDIT_XLSX = os.getenv("AUDIT_XLSX", "eval/eval_hybrid_rval2.xlsx")
RUN_ID = datetime.now().strftime("%Y%m%d-%H%M%S")
RUN_TS = datetime.now().isoformat(timespec="seconds")

def _append_df_to_excel(path: str, sheet_name: str, df: pd.DataFrame):
    """
    Append DataFrame ke Excel. Buat file/sheet bila belum ada.
    Kompatibel dengan pandas 2.x (tanpa set writer.book).
    """
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    # Jika file belum ada → buat baru
    if not os.path.exists(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # File sudah ada → tentukan startrow (kalau sheet ada)
    try:
        book = load_workbook(path)
        if sheet_name in book.sheetnames:
            ws = book[sheet_name]
            startrow = ws.max_row or 0
            # Jika baris pertama kosong (sheet baru tapi ada row default), reset ke 0
            if ws.max_row == 1:
                first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                if all(v is None for v in first_row):
                    startrow = 0
        else:
            startrow = 0
        book.close()
    except Exception:
        # Jika gagal baca workbook (misal corrupt/locked), fallback ke tulis dari awal sheet
        startrow = 0

    # Append tanpa menyentuh writer.book (biarkan pandas yang handle)
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(startrow == 0),
            startrow=startrow
        )


# ===============================
# 1) ENV & API CONFIG
# ===============================
DATA_API_URL = os.getenv("DATA_API_URL", "").rstrip("/")
API_TOKEN = os.getenv("API_TOKEN", "")
QUERY_PATH = "/query_listing"
TIMEOUT = 15
XLS_FILE = "test/test_hybrid_rval1.xlsx"

# ===============================
# 2) UTIL: NORMALISASI & PARSERS
# ===============================
def normalize_ws(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()

def parse_int_from_money(s: str):
    nums = re.sub(r"[^\d]", "", s)
    return int(nums) if nums else None

def extract_price(text: str):
    m = re.search(r"[Rr]p\s*[\d\.\,]+", text)
    return parse_int_from_money(m.group(0)) if m else None

def extract_bedrooms(text: str):
    patterns = [
        r"[Kk]amar\s*[Tt]idur[:\s]*([0-9]+)",
        r"\b([0-9]+)\s*[Kk]amar\s*[Tt]idur\b",
        r"\bKT[:\s-]*([0-9]+)\b",
        r"\b([0-9]+)\s*KT\b",
        r"\b([0-9]+)\s*(BR|Bed|Bedroom)s?\b",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.I)
        if m: return int(m.group(1))
    return None

def extract_luas_bangunan(text: str):
    patterns = [
        r"\bLB[:\s-]*([0-9]+)\s*m?\s*2?\b",
        r"[Ll]uas\s*[Bb]angunan[:\s-]*([0-9]+)\s*m",
        r"[Bb]uilding\s*[Aa]rea[:\s-]*([0-9]+)\s*m",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m: return int(m.group(1))
    return None

def extract_luas_tanah(text: str):
    patterns = [
        r"\bLT[:\s-]*([0-9]+)\s*m?\s*2?\b",
        r"[Ll]uas\s*[Tt]anah[:\s-]*([0-9]+)\s*m",
        r"[Ll]and\s*[Aa]rea[:\s-]*([0-9]+)\s*m",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m: return int(m.group(1))
    return None

def extract_lebar_bangunan(text: str):
    patterns = [
        r"[Ll]ebar(?:\s*[Bb]angunan)?[:\s-]*([0-9]+(?:\.[0-9]+)?)\s*m?\b",
        r"\b([0-9]+(?:\.[0-9]+)?)\s*x\s*[0-9]+(?:\.[0-9]+)?\b",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            try:
                return float(m.group(1))
            except:
                pass
    return None

def extract_jumlah_tingkat(text: str):
    patterns = [
        r"[Jj]umlah\s*[Tt]ingkat[:\s-]*([0-9]+)",
        r"\b([0-9]+)\s*[Tt]ingkat\b",
        r"\b([0-9]+)\s*[Ll]antai\b",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m: return int(m.group(1))
    return None

def extract_kondisi(text: str):
    t = text.lower()
    mapping = {
        "baru": ["baru", "brand new", "new"],
        "kosong": ["kosong", "unfurnished", "tanpa perabot", "empty"],
        "full furnished": ["full furnished", "fully furnished", "lengkap perabot", "fullfurnished"],
        "non furnished": ["non furnished", "semi furnished", "partial furnished", "semi-furnished"],
    }
    for k, alts in mapping.items():
        for a in alts:
            if re.search(r"\b" + re.escape(a) + r"\b", t):
                return k
    return None

def infer_jenis_properti(text: str) -> int | None:
    t = text.lower()

    # 1) sinyal kuat by phrases (judul/teks)
    def has(pat): 
        return re.search(pat, t, flags=re.I) is not None

    # --- RUKO (2) ---
    if has(r"\bruko\b") or has(r"\brumah\s*toko\b") or has(r"\bruko\s+gudang\b"):
        return 2

    # --- APARTEMEN (4) ---
    if has(r"\bapart(e)?men(t)?\b") or has(r"\bapartment\b") or has(r"\bcondo(minium)?\b"):
        return 4

    # --- GUDANG (5) ---
    if has(r"\bgudang\b") or has(r"\bwarehouse\b"):
        return 5

    # --- GEDUNG / PERKANTORAN (6) ---
    if has(r"\bgedung\b") or has(r"\bperkantoran\b") or has(r"\boffice\s+building\b"):
        return 6

    # --- TANAH (3) dengan konteks "lahan/kavling/tanah dijual"
    tanah_context = (
        has(r"\btanah\s*dijual\b") or
        has(r"\blahan\b") or
        has(r"\bkav?ling\b") or
        has(r"\btanah\s*kapling\b") or
        (has(r"\btanah\b") and not has(r"(luas\s*tanah|lt[:\s-]*\d|sertifikat\s*tanah|shm|m2)"))
    )
    if tanah_context:
        return 3

    # --- RUMAH (1) ---
    if has(r"\brumah\b") or has(r"\bhouse\b"):
        return 1

    return None


def infer_tipe_listing(text: str):
    t = text.lower()
    if re.search(r"\blelang\b|\bauction\b", t):
        return 3
    if re.search(r"\bdijual\b|\bfor sale\b", t):
        return 1
    if re.search(r"\bdisewa(?:kan)?\b|\bsewa\b|\bfor rent\b|\brent\b|\bkontrakan\b|\bdikontrakk?an\b", t):
        return 2
    return None

def extract_mata_angin(text: str):
    t = text.lower()
    arah_map = {
        "utara": ["utara", "north", r"\bn\b"],
        "timur laut": ["timur laut", "timur-laut", "timurlaut", "northeast", r"\bne\b"],
        "timur": ["timur", "east", r"\be\b"],
        "tenggara": ["tenggara", "southeast", r"\bse\b"],
        "selatan": ["selatan", "south", r"\bs\b"],
        "barat daya": ["barat daya", "barat-daya", "baratdaya", "southwest", r"\bsw\b"],
        "barat": ["barat", "west", r"\bw\b"],
        "barat laut": ["barat laut", "barat-laut", "baratlaut", "northwest", r"\bnw\b"],
    }
    ctx = t
    m_ctx = re.search(r"(hadap|menghadap|facing|orientasi)[:\s]*([a-z\- ]+)", t)
    if m_ctx:
        start = m_ctx.start()
        ctx = t[start:start+60]
    for canon, alts in arah_map.items():
        for a in alts:
            if re.search(r"\b" + a + r"\b", ctx):
                return canon
    return None

# ---------- Keyword matching (tanpa hardcode) ----------
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def _normalize(s: str) -> str:
    s = _strip_accents(s.lower())
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _normalize_for_words(s: str) -> str:
    s = _strip_accents(s.lower())
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _word_in(text: str, kw: str) -> bool:
    if not text or not kw:
        return False
    t = _normalize_for_words(text)
    k = _normalize_for_words(kw)
    return re.search(rf"\b{re.escape(k)}\b", t) is not None

def _generate_variants(kw: str):
    base = _normalize(kw)
    variants = set()
    variants.add(base)
    variants.add(base.replace(" ", "-"))
    variants.add(base.replace("-", " "))
    variants.add(base.replace(" ", "").replace("-", ""))  # tanpa spasi & hyphen
    return variants

def contains_keyword(text: str, kw: str, aliases: list[str] | None = None, threshold: int = 85) -> bool:
    """Cek keyword di text secara fuzzy (tanpa hardcode), dukung alias dari API."""
    if not kw:
        return True
    t = _normalize(text)
    candidates = set([kw])
    if aliases:
        for a in aliases:
            if a and a.strip():
                candidates.add(a)
    all_variants = set()
    for c in candidates:
        all_variants.update(_generate_variants(c))
    for v in all_variants:
        if fuzz.token_set_ratio(t, v) >= threshold or fuzz.partial_ratio(t, v) >= threshold:
            return True
    return False

def contains_phrases(text: str, phrases, threshold: int = 85, mode: str = "all") -> bool:
    """
    Fuzzy-check apakah 'text' mengandung frasa2 pada 'phrases'.
    - phrases: string dipisah koma/semicolon/pipe ATAU list[str]
    - threshold: skor RapidFuzz minimal (token_set_ratio / partial_ratio)
    - mode: "all" (default) = semua frasa harus match, "any" = salah satu saja
    """
    if not text:
        return False

    # Normalisasi frasa -> list
    if isinstance(phrases, str):
        phrases_list = [p.strip() for p in re.split(r"[,;|]", phrases) if p.strip()]
    elif isinstance(phrases, (list, tuple, set)):
        phrases_list = [str(p).strip() for p in phrases if str(p).strip()]
    else:
        return False

    if not phrases_list:
        return False

    base = _normalize(text)
    hits = 0
    for ph in phrases_list:
        phn = _normalize(ph)
        score = max(fuzz.token_set_ratio(base, phn), fuzz.partial_ratio(base, phn))
        ok = score >= threshold
        if mode == "any" and ok:
            return True
        if ok:
            hits += 1

    return (hits == len(phrases_list)) if mode == "all" else False

def _collect_extra_text_from_truth(truth_obj: dict, include_title: bool = True) -> str:
    """
    Ambil teks untuk evaluasi 'info_lainnya' dari objek API:
    - info_tambahan (string / list)
    - + judul (opsional) bila include_title=True

    Mengembalikan gabungan teks (dipisah newline) atau string kosong.
    """
    if not isinstance(truth_obj, dict):
        return ""

    # Pastikan resolve dulu
    truth = _resolve_truth_obj(truth_obj) or {}

    pieces = []

    # 1) info_tambahan (prioritas utama)
    val = truth.get("info_tambahan")
    if isinstance(val, str) and val.strip():
        pieces.append(val.strip())
    elif isinstance(val, (list, tuple)):
        joined = " \n ".join([str(v).strip() for v in val if str(v).strip()])
        if joined:
            pieces.append(joined)

    # 2) judul (opsional) — cek beberapa kemungkinan nama kunci
    if include_title:
        for k in ("judul", "title", "nama", "judul_listing"):
            v = truth.get(k)
            if isinstance(v, str) and v.strip():
                pieces.append(v.strip())
                break  # ambil satu yang pertama ketemu

    return " \n ".join(pieces).strip()




def extract_title_snippet(text: str) -> str:
    m = re.search(r"\*(.+?)\*", text)
    if m:
        t = m.group(1)
    else:
        first = next((ln.strip() for ln in text.splitlines() if ln.strip()), "")
        t = first[:80]
    return re.sub(r"\s+", " ", t).strip()

def extract_first_link(text: str) -> str | None:
    m = re.search(r"https?://\S+", text)
    return m.group(0) if m else None

def parse_listing_id_from_link(link: str) -> int | None:
    m = re.search(r"/listing/(\d+)", link)
    if m:
        try: return int(m.group(1))
        except: return None
    m2 = re.search(r"/(\d+)(?:\D|$)", link)
    if m2:
        try: return int(m2.group(1))
        except: return None
    return None

# ===============================
# 3) SPLITTER
# ===============================
def split_listings(raw_text: str):
    text = raw_text.strip("\n")
    links = re.findall(r"https?://\S+", text)
    if len(links) == 1:
        return [text]
    starts = [m.start() for m in re.finditer(r"(?m)^\s*(\d+[\.\)]\s+|\*(?!\*)\s+)", text)]
    items = []
    if len(starts) >= 2:
        for i, s in enumerate(starts):
            e = starts[i+1] if i+1 < len(starts) else len(text)
            chunk = text[s:e].strip()
            if len(chunk) > 10:
                items.append(chunk)
        if items and sum(len(c) < 120 for c in items) >= len(items) - 1:
            return [text]
        return items
    parts = re.split(r"(?i)(?:^|\n)\s*-\s*\*\*Link:\*\*|\n\s*Link:", text)
    if len(parts) > 1:
        rebuilt = []
        for i in range(1, len(parts)):
            segment = parts[i-1].split("\n\n")[-1] + "\nLink:" + parts[i].split("\n\n")[0]
            segment = segment.strip()
            if len(segment) > 10:
                rebuilt.append(segment)
        if rebuilt:
            return rebuilt
    return [text.strip()]

# ===============================
# 4) API CALL (GROUND TRUTH)
# ===============================
_session_cache = {}

def fetch_truth_from_api(listing_id: int) -> dict | None:
    if not listing_id or not DATA_API_URL or not API_TOKEN:
        return None
    if listing_id in _session_cache:
        return _session_cache[listing_id]

    url = f"{DATA_API_URL}{QUERY_PATH}"
    headers = {
        "Authorization": f"Bearer {API_TOKEN}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    payload = {"listing_id": int(listing_id)}
    # print(payload)
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        # print(data)
        _session_cache[listing_id] = data
        # print(data)
        return data
    except requests.exceptions.Timeout:
        print(f"[API] Timeout listing_id={listing_id}")
    except requests.exceptions.HTTPError as e:
        print(f"[API] HTTP error: {e} | body: {resp.text if 'resp' in locals() else ''}")
    except requests.exceptions.RequestException as e:
        print(f"[API] Request error: {e}")
    except Exception as e:
        print(f"[API] Unexpected error: {e}")
    return None

def _resolve_truth_obj(truth):
    if not isinstance(truth, dict):
        return None
    if any(k in truth for k in ("alamat_ditampilkan", "keyword", "total_harga_listing", "k_tidur")):
        return truth
    data = truth.get("data")
    if isinstance(data, dict):
        return data
    if isinstance(data, list) and data:
        first = data[0]
        if isinstance(first, dict):
            return first
    listing = truth.get("listing")
    if isinstance(listing, dict):
        return listing
    return truth

def merge_extracted_with_truth(extracted: dict, truth: dict | None) -> dict:
    if not truth:
        return extracted.copy()
    merged = extracted.copy()
    truth_resolved = _resolve_truth_obj(truth)
    if truth_resolved.get("total_harga_listing") is not None:
        merged["price"] = int(truth_resolved["total_harga_listing"])
    if truth_resolved.get("k_tidur") is not None:
        merged["kamar_tidur"] = int(truth_resolved["k_tidur"])
    if truth_resolved.get("bangunan_luas") is not None:
        merged["luas_bangunan"] = int(truth_resolved["bangunan_luas"])
    if truth_resolved.get("tanah_luas") is not None:
        merged["luas_tanah"] = int(truth_resolved["tanah_luas"])
    if truth_resolved.get("bangunan_tingkat") is not None:
        merged["jumlah_tingkat"] = int(truth_resolved["bangunan_tingkat"])
    if truth_resolved.get("jenis_properti") is not None:
        merged["jenis_properti"] = int(truth_resolved["jenis_properti"])
    if truth_resolved.get("tipe_listing") is not None:
        merged["tipe_listing"] = int(truth_resolved["tipe_listing"])
    if truth_resolved.get("arah_site"):
        merged["mata_angin"] = str(truth_resolved["arah_site"]).lower().strip()
    merged["_truth_keywords"] = truth_resolved.get("keyword", "")
    merged["_truth_obj"] = truth_resolved
    return merged

# ===============================
# 5) KEYWORD MATCH (with alamat)
# ===============================
def keyword_match_with_truth(answer_text: str, gold_kw: str, truth_keywords_or_obj, threshold=85) -> bool:
    if not gold_kw:
        return True
    if contains_keyword(answer_text, gold_kw, threshold=threshold):
        return True
    aliases = []
    alamat_api = None
    if isinstance(truth_keywords_or_obj, dict):
        truth_obj = _resolve_truth_obj(truth_keywords_or_obj)
        tk = truth_obj.get("keyword")
        if isinstance(tk, str):
            aliases.extend([t.strip() for t in re.split(r"[,;|]", tk) if t.strip() ])
        elif isinstance(tk, list):
            aliases.extend([str(t).strip() for t in tk if str(t).strip()])
        alamat_api = truth_obj.get("alamat_ditampilkan")
    elif isinstance(truth_keywords_or_obj, str):
        aliases.extend([t.strip() for t in re.split(r"[,;|]", truth_keywords_or_obj) if t.strip()])
    elif isinstance(truth_keywords_or_obj, list):
        aliases.extend([str(t).strip() for t in truth_keywords_or_obj if str(t).strip()])
    if aliases:
        g = _normalize(gold_kw)
        for a in aliases:
            a_norm = _normalize(a)
            if (fuzz.token_set_ratio(g, a_norm) >= threshold) or (fuzz.partial_ratio(g, a_norm) >= threshold):
                return True
    if isinstance(alamat_api, str) and alamat_api.strip():
        if _word_in(alamat_api, gold_kw):
            return True
        if contains_keyword(alamat_api, gold_kw, threshold=70):
            return True
    return False

# ===============================
# 6) EKSTRAK & EVALUASI PER ITEM
# ===============================
def extract_all_fields(text: str):
    t = normalize_ws(text)
    return {
        "price": extract_price(t),
        "kamar_tidur": extract_bedrooms(t),
        "luas_bangunan": extract_luas_bangunan(t),
        "luas_tanah": extract_luas_tanah(t),
        "lebar_bangunan": extract_lebar_bangunan(t),
        "jumlah_tingkat": extract_jumlah_tingkat(t),
        "kondisi": extract_kondisi(t),
        "jenis_properti": infer_jenis_properti(t),
        "tipe_listing": infer_tipe_listing(t),
        "mata_angin": extract_mata_angin(t),
    }

def evaluate_constraints(gold: dict, extracted: dict, answer_text: str):
    if "luar_bangunan" in gold and "luas_bangunan" not in gold:
        gold = {**gold, "luas_bangunan": gold["luar_bangunan"]}
    preds = {}
    def has(k):
        return k in gold and gold[k] not in (None, "", [])
    if has("keyword"):
        preds["keyword"] = keyword_match_with_truth(
            answer_text,
            gold["keyword"],
            extracted.get("_truth_obj") if extracted.get("_truth_obj") is not None else extracted.get("_truth_keywords")
        )
    if has("harga_min"):
        preds["harga_min"] = (extracted["price"] is not None and extracted["price"] >= gold["harga_min"])
    if has("harga_max"):
        preds["harga_max"] = (extracted["price"] is not None and extracted["price"] <= gold["harga_max"])
    if has("kamar_tidur"):
        preds["kamar_tidur"] = (extracted["kamar_tidur"] is not None and extracted["kamar_tidur"] >= gold["kamar_tidur"])
    if has("lebar_bangunan"):
        preds["lebar_bangunan"] = (extracted["lebar_bangunan"] is not None and extracted["lebar_bangunan"] >= gold["lebar_bangunan"])
    if has("luas_bangunan"):
        preds["luas_bangunan"] = (extracted["luas_bangunan"] is not None and extracted["luas_bangunan"] >= gold["luas_bangunan"])
    if has("jumlah_tingkat"):
        preds["jumlah_tingkat"] = (extracted["jumlah_tingkat"] is not None and extracted["jumlah_tingkat"] == gold["jumlah_tingkat"])
    if has("luas_tanah"):
        preds["luas_tanah"] = (extracted["luas_tanah"] is not None and extracted["luas_tanah"] >= gold["luas_tanah"])
    if has("kondisi"):
        gk = gold["kondisi"].lower().strip()
        preds["kondisi"] = (extracted["kondisi"] is not None and extracted["kondisi"] == gk)
    if has("jenis_properti"):
        preds["jenis_properti"] = (extracted["jenis_properti"] == gold["jenis_properti"])
    if has("tipe_listing"):
        preds["tipe_listing"] = (extracted["tipe_listing"] == gold["tipe_listing"])
    if has("mata_angin"):
        ga = str(gold["mata_angin"]).lower().strip()
        preds["mata_angin"] = (extracted["mata_angin"] is not None and extracted["mata_angin"] == ga)
        # --- info_lainnya: cek di teks 'Keterangan Tambahan' dari API ---
    if has("info_lainnya"):
        truth_obj = extracted.get("_truth_obj")
        extra_text = ""
        if truth_obj:
            truth_obj = _resolve_truth_obj(truth_obj) or {}
            extra_text = _collect_extra_text_from_truth(truth_obj)

        phrases = gold["info_lainnya"]
        # Mode default: "all" (semua frasa harus ada). Jika mau longgar, ganti ke "any".
        preds["info_lainnya"] = contains_phrases(extra_text, phrases, threshold=80, mode="all") if extra_text else False

    return preds

# ===============================
# 7) NO-RESULT DETECTOR & API CHECK
# ===============================
def looks_like_no_result_answer(text: str) -> bool:
    t = text.lower()
    patterns = [
        r"\bmaaf\b.*\btidak (menemukan|ada)\b",
        r"\btidak (menemukan|ada)\b.*\b(data|hasil)\b",
        r"\bbelum ada\b.*\b(data|hasil)\b",
        r"\bno (result|results)\b",
        r"\bno listings?\b",
        r"\bnot found\b",
        r"\bdata (tidak|nggak|gak) ditemukan\b",
        r"\bhasil pencarian\b.*\bkosong\b",
        r"\bbelum tersedia\b",
        r"\btidak ada\b.*\b(rumah|listing|properti|unit)\b",
        r"\bsemua\s+listing\b.*\b(di atas|lebih dari|>\s*)\b",
        r"\bbelum ditemukan\b",
        r"\btidak tersedia\b.*\b(rumah|listing|properti|unit)\b",
    ]
    if any(re.search(p, t) for p in patterns):
        return True
    has_link = bool(extract_first_link(text))
    has_price = bool(extract_price(text))
    has_bed = bool(extract_bedrooms(text))
    has_any_number = bool(re.search(r"\d", text))
    negative_tone = bool(re.search(r"\b(tidak|nggak|gak|bukan)\b", t))
    if (not has_link) and (not has_price) and (not has_bed) and negative_tone and (not has_any_number):
        return True
    return False

def listings_empty_for_gold(gold: dict) -> tuple[bool, str]:
    if not DATA_API_URL or not API_TOKEN:
        return (None, "API check skipped: missing DATA_API_URL/API_TOKEN")
    url = f"{DATA_API_URL}{QUERY_PATH}"
    headers = {"Authorization": f"Bearer {API_TOKEN}", "Accept": "application/json"}
    allow_keys = {
        "keyword", "harga_min", "harga_max", "kamar_tidur",
        "jenis_properti", "tipe_listing", "jumlah_tingkat",
        "luas_bangunan", "luas_tanah", "lebar_bangunan", "mata_angin",
        "alamat", "page", "paginate"
    }
    payload = {k: v for k, v in gold.items() if k in allow_keys}
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=TIMEOUT)
        resp.raise_for_status()
        data = resp.text  # tetap logic lama
        def _is_empty(x):
            if x is None: return True
            if x == "": return True
            if isinstance(x, (list, tuple, set)) and len(x) == 0: return True
            if isinstance(x, dict):
                if x.get("data") == []: return True
                if x.get("rows") == []: return True
                if x.get("count") in (0, "0"): return True
                if "data" in x and isinstance(x["data"], list) and len(x["data"]) == 0:
                    return True
                if len(x) == 0: return True
            return False
        empty = _is_empty(data)
        reason = "API empty" if empty else "API has data"
        return (empty, reason)
    except requests.exceptions.Timeout:
        return (None, "API timeout")
    except requests.exceptions.HTTPError as e:
        return (None, f"API HTTP error: {e}")
    except requests.exceptions.RequestException as e:
        return (None, f"API request error: {e}")
    except Exception as e:
        return (None, f"API unexpected error: {e}")

# ===============================
# 8) LOOPING SEMUA ITEM & METRIK
# ===============================
def split_and_evaluate(answer_text: str, gold: dict):
    nr_claim = looks_like_no_result_answer(answer_text)
    if nr_claim:
        is_empty, api_note = listings_empty_for_gold(gold)
        print("== NO-RESULT MODE ==")
        print(f"Detected no-result claim from answer: {nr_claim}")
        print(f"API check: {api_note}")
        if is_empty is True:
            print("No-result correctness: TRUE (jawaban sesuai: API juga kosong)")
            no_result_score = 1.0
        elif is_empty is False:
            print("No-result correctness: FALSE (jawaban salah: API ada data)")
            no_result_score = 0.0
        else:
            print("No-result correctness: UNKNOWN (API check tidak konklusif)")
            no_result_score = None
        print("\n== METRICS (No-listing) ==")
        print("CPR@All: N/A (0 items)")
        print("CPR@5  : N/A (0 items)")
        print("Avg Per-Constraint Accuracy: "
              + (f"{no_result_score:.3f}" if no_result_score is not None else "N/A"))
        print("Precision / Recall / F1: N/A (0 items)")
        return {
            "has_items": False,
            "avg_pca": no_result_score,
            "precision": None, "recall": None, "f1": None,
            "cpr_all": None, "cpr_at_5": None,
            "constraints_total": 0,
            "constraints_correct": 0,
            "no_result_score": no_result_score,
            "items": [],
            "items_count": 0,
            "api_note": api_note,
            "no_result_claim": True
        }
    items = split_listings(answer_text)
    print(f"Total items (detected): {len(items)}")
    for i, it in enumerate(items, 1):
        if len(it) < 120:
            print(f"    [WARN] Item #{i} sangat pendek—cek hasil split.")
    all_preds = []
    all_y_true = []
    all_y_pred = []
    strict_successes = 0
    for idx, item in enumerate(items, start=1):
        title = extract_title_snippet(item)
        link = extract_first_link(item)
        listing_id = parse_listing_id_from_link(link) if link else None
        ext = extract_all_fields(item)
        truth = fetch_truth_from_api(listing_id) if listing_id else None
        ext_merged = merge_extracted_with_truth(ext, truth)
        pred = evaluate_constraints(gold, ext_merged, item)
        keys = list(pred.keys())
        y_true = [1] * len(keys)
        y_pred = [int(bool(pred[k])) for k in keys]
        per_constraint_accuracy = (sum(y_pred)/len(y_pred) if y_pred else 0.0)
        strict = int(all(y_pred) if y_pred else False)
        all_preds.append({
            "index": idx,
            "title": title,
            "link": link,
            "listing_id": listing_id,
            "extracted": ext_merged,
            "pred": pred,
            "per_constraint_accuracy": per_constraint_accuracy,
            "strict_success": strict
        })
        strict_successes += strict
        all_y_true.extend(y_true)
        all_y_pred.extend(y_pred)
    k_all = len(items)
    cpr_all = strict_successes / k_all if k_all else 0.0
    k5 = min(5, k_all)
    cpr_at_5 = (sum(all_preds[i]["strict_success"] for i in range(k5)) / k5) if k5 else 0.0
    if all_y_true:
        prec, rec, f1, _ = precision_recall_fscore_support(
            all_y_true, all_y_pred, average="binary", zero_division=0
        )
        micro_correct = sum(all_y_pred)
        micro_total = len(all_y_pred)
    else:
        prec = rec = f1 = 0.0
        micro_correct = 0
        micro_total = 0
    avg_pca = sum(d["per_constraint_accuracy"] for d in all_preds) / len(all_preds) if all_preds else 0.0
    for d in all_preds:
        print("Pred:", d["pred"])
        print("Per-constraint accuracy:", round(d["per_constraint_accuracy"], 3))
        print("Strict success:", d["strict_success"])
    print("\n== METRICS ==")
    print(f"CPR@All: {cpr_all:.3f}")
    print(f"CPR@5  : {cpr_at_5:.3f}")
    print(f"Avg Per-Constraint Accuracy: {avg_pca:.3f}")
    print(f"Precision: {prec:.3f} | Recall: {rec:.3f} | F1: {f1:.3f}")
    return {
        "has_items": True,
        "avg_pca": avg_pca, "precision": prec, "recall": rec, "f1": f1,
        "cpr_all": cpr_all, "cpr_at_5": cpr_at_5,
        "constraints_total": micro_total, "constraints_correct": micro_correct,
        "no_result_score": None,
        "items": all_preds,
        "items_count": len(all_preds),
        "api_note": None,
        "no_result_claim": False
    }

# ===============================
# 8.5) CM HELPER
# ===============================
def _gt_from_api(gold: dict):
    is_empty, note = listings_empty_for_gold(gold)
    if is_empty is True:
        return False, note
    elif is_empty is False:
        return True, note
    else:
        return None, note

def classify_confusion_from_summary(gold: dict, summary: dict, T: float):
    gt_pos, gt_note = _gt_from_api(gold)
    has_items = bool(summary.get("has_items", False))
    nores_conclusive = (not has_items) and (summary.get("no_result_score") is not None)
    cpr_all = summary.get("cpr_all", 0.0) or 0.0
    pred_negative = nores_conclusive
    pred_positive = (has_items and (cpr_all >= T))
    if gt_pos is None:
        label = "UNK"
    else:
        if gt_pos:   # GT+
            label = "TP" if pred_positive else "FN"
        else:        # GT-
            label = "TN" if pred_negative else "FP"
    info = {
        "gt": ("Pos" if gt_pos is True else ("Neg" if gt_pos is False else "Unknown")),
        "gt_note": gt_note,
        "has_items": has_items,
        "no_result": nores_conclusive,
        "cpr_all": cpr_all,
        "threshold": T,
        "pred": ("Pos" if pred_positive else "Neg"),
    }
    return label, info

# ===============================
# 9) JALANKAN + APPEND KE EXCEL
# ===============================
if __name__ == "__main__":
    excel_file = XLS_FILE
    result = build_answer_gold_list(excel_file, sheet_name=0)
    T = float(os.getenv("CPR_THRESHOLD", "0.60"))

    # aggregator
    n_questions = 0
    n_with_items = 0
    macro_pca_sum = 0.0
    macro_p_sum = macro_r_sum = macro_f1_sum = 0.0
    macro_p_count = macro_r_count = macro_f1_count = 0
    micro_correct_total = 0
    micro_constraints_total = 0
    no_result_scores = []
    n_for_macro_pca = 0

    TP = FP = FN = TN = UNK = 0

    # buffers untuk Excel
    per_q_rows = []
    per_item_rows = []

    for i, item in enumerate(result, start=1):
        print(f"== QUESTION #{i} ==")
        answer_text = item["answer"]
        gold_str = item["gold"]
        gold = ast.literal_eval(gold_str)

        # excerpt untuk audit
        answer_excerpt = normalize_ws(answer_text)[:200]

        summary = split_and_evaluate(answer_text, gold)

        # akumulasi (logic lama)
        n_questions += 1
        if summary["avg_pca"] is not None:
            macro_pca_sum += summary["avg_pca"]; n_for_macro_pca += 1
        if summary["has_items"]:
            n_with_items += 1
            macro_p_sum += (summary["precision"] or 0.0)
            macro_r_sum += (summary["recall"] or 0.0)
            macro_f1_sum += (summary["f1"] or 0.0)
            macro_p_count += 1; macro_r_count += 1; macro_f1_count += 1
            micro_correct_total += (summary["constraints_correct"] or 0)
            micro_constraints_total += (summary["constraints_total"] or 0)
        else:
            if summary["no_result_score"] is not None:
                no_result_scores.append(summary["no_result_score"])

        # Confusion Matrix per-pertanyaan
        label, info = classify_confusion_from_summary(gold, summary, T)
        if   label == "TP": TP += 1
        elif label == "FP": FP += 1
        elif label == "FN": FN += 1
        elif label == "TN": TN += 1
        else: UNK += 1

        print("[CM] Label={label} | GT={gt} | Pred={pred} | "
              "has_items={has_items} | no_result={nores} | "
              "CPR@All={cpr:.3f} (T={T:.2f}) | API='{note}'"
              .format(
                  label=label, gt=info["gt"], pred=info["pred"],
                  has_items=info["has_items"], nores=info["no_result"],
                  cpr=info["cpr_all"], T=info["threshold"], note=info["gt_note"]
              ))
        print("\n\n")

        # ====== APPEND ROW: per_question ======
        per_q_rows.append({
            "run_id": RUN_ID,
            "run_ts": RUN_TS,
            "xls_input": excel_file,
            "q_index": i,
            "has_items": summary["has_items"],
            "no_result_claim": summary.get("no_result_claim"),
            "api_note": summary.get("api_note"),
            "items_count": summary.get("items_count"),
            "avg_pca": summary["avg_pca"],
            "precision": summary["precision"],
            "recall": summary["recall"],
            "f1": summary["f1"],
            "cpr_all": summary["cpr_all"],
            "cpr_at_5": summary["cpr_at_5"],
            "constraints_total": summary["constraints_total"],
            "constraints_correct": summary["constraints_correct"],
            "no_result_score": summary["no_result_score"],
            "cm_label": label,
            "cm_gt": info["gt"],
            "cm_pred": info["pred"],
            "cm_threshold": info["threshold"],
            "answer_excerpt": answer_excerpt,
            "gold_json": json.dumps(gold, ensure_ascii=False)
        })

        # ====== APPEND ROWS: per_item ======
        for it in summary.get("items", []):
            ext = it.get("extracted", {}) or {}
            pred_map = it.get("pred", {}) or {}
            per_item_rows.append({
                "run_id": RUN_ID,
                "run_ts": RUN_TS,
                "xls_input": excel_file,
                "q_index": i,
                "item_idx": it.get("index"),
                "title": it.get("title"),
                "link": it.get("link"),
                "listing_id": it.get("listing_id"),
                "per_constraint_accuracy": it.get("per_constraint_accuracy"),
                "strict_success": it.get("strict_success"),
                # ringkasan hasil ekstraksi (pasca-override truth)
                "price": ext.get("price"),
                "kamar_tidur": ext.get("kamar_tidur"),
                "luas_bangunan": ext.get("luas_bangunan"),
                "luas_tanah": ext.get("luas_tanah"),
                "lebar_bangunan": ext.get("lebar_bangunan"),
                "jumlah_tingkat": ext.get("jumlah_tingkat"),
                "kondisi": ext.get("kondisi"),
                "jenis_properti": ext.get("jenis_properti"),
                "tipe_listing": ext.get("tipe_listing"),
                "mata_angin": ext.get("mata_angin"),
                # simpan JSON predikat utk audit
                "pred_json": json.dumps(pred_map, ensure_ascii=False)
            })

    # hitung agregat (logic lama)
    macro_avg_pca = (macro_pca_sum / n_for_macro_pca) if n_for_macro_pca else None
    macro_avg_p = (macro_p_sum / macro_p_count) if macro_p_count else None
    macro_avg_r = (macro_r_sum / macro_r_count) if macro_r_count else None
    macro_avg_f1 = (macro_f1_sum / macro_f1_count) if macro_f1_count else None
    micro_accuracy = (micro_correct_total / micro_constraints_total) if micro_constraints_total else None
    avg_no_result_score = (sum(no_result_scores) / len(no_result_scores)) if no_result_scores else None

    print("======== SUMMARY (GLOBAL) ========")
    print(f"Total questions                 : {n_questions}")
    print(f"Total with data (has items)     : {n_with_items}")
    print(f"Total no-result (conclusive)    : {len(no_result_scores)}")
    print(f"Macro Avg Per-Constraint Acc    : {macro_avg_pca:.3f}" if macro_avg_pca is not None else "Macro Avg Per-Constraint Acc    : N/A")
    print(f"Macro Avg Precision             : {macro_avg_p:.3f}" if macro_avg_p is not None else "Macro Avg Precision             : N/A")
    print(f"Macro Avg Recall                : {macro_avg_r:.3f}" if macro_avg_r is not None else "Macro Avg Recall                : N/A")
    print(f"Macro Avg F1                    : {macro_avg_f1:.3f}" if macro_avg_f1 is not None else "Macro Avg F1                    : N/A")
    print(f"Micro Accuracy (pooled)         : {micro_accuracy:.3f}" if micro_accuracy is not None else "Micro Accuracy (pooled)         : N/A")
    print(f"Avg No-Result Score             : {avg_no_result_score:.3f}" if avg_no_result_score is not None else "Avg No-Result Score             : N/A")
    print(f"Threshold T (CPR@All)           : {T:.2f}")
    print(f"Confusion Matrix                : TP={TP} | FP={FP} | FN={FN} | TN={TN} | UNK={UNK}")

    cm_precision = (TP / (TP + FP)) if (TP + FP) > 0 else 0.0
    cm_recall    = (TP / (TP + FN)) if (TP + FN) > 0 else 0.0
    cm_f1        = (2 * cm_precision * cm_recall / (cm_precision + cm_recall)) if (cm_precision + cm_recall) > 0 else 0.0
    cm_accuracy  = ((TP + TN) / (TP + TN + FP + FN)) if (TP + TN + FP + FN) > 0 else 0.0

    print("---- Confusion Matrix Metrics ----")
    print(f"CM Precision                    : {cm_precision:.3f}")
    print(f"CM Recall                       : {cm_recall:.3f}")
    print(f"CM F1                           : {cm_f1:.3f}")
    print(f"CM Accuracy                     : {cm_accuracy:.3f}")

    # ========= TULIS KE EXCEL (APPEND) =========
    if per_q_rows:
        _append_df_to_excel(AUDIT_XLSX, "per_question", pd.DataFrame(per_q_rows))
    if per_item_rows:
        _append_df_to_excel(AUDIT_XLSX, "per_item", pd.DataFrame(per_item_rows))

    global_row = pd.DataFrame([{
        "run_id": RUN_ID,
        "run_ts": RUN_TS,
        "xls_input": XLS_FILE,
        "threshold_T": T,
        "total_questions": n_questions,
        "total_with_items": n_with_items,
        "total_no_result_conclusive": len(no_result_scores),
        "macro_avg_pca": macro_avg_pca,
        "macro_avg_precision": macro_avg_p,
        "macro_avg_recall": macro_avg_r,
        "macro_avg_f1": macro_avg_f1,
        "micro_accuracy": micro_accuracy,
        "avg_no_result_score": avg_no_result_score,
        "TP": TP, "FP": FP, "FN": FN, "TN": TN, "UNK": UNK,
        "cm_precision": cm_precision,
        "cm_recall": cm_recall,
        "cm_f1": cm_f1,
        "cm_accuracy": cm_accuracy
    }])
    _append_df_to_excel(AUDIT_XLSX, "global_runs", global_row)

    print(f"[AUDIT] Appended to Excel: {AUDIT_XLSX}")
